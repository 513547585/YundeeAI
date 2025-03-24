# coding=utf8
import random

import requests
import xlrd
from openpyxl import load_workbook


def file_ID():
    data_excel = xlrd.open_workbook('case.xlsx')
    names = data_excel.sheet_names()  # 获取所有sheet名称
    print("sheet名称合集：", names)
    # print('all sheets --- {}'.format(names))
    table1 = data_excel.sheets()[0]  # 通过索引顺序获取sheet
    # 获取文件的sheet 行数 列数
    sheet_name = table1.name
    rows_number = table1.nrows
    cols_number = table1.ncols
    print(sheet_name)  # sheet
    print("sheet行数：", rows_number)  # 行数
    print("sheet列数：", cols_number)  # 列数
    row_value = table1.row_values(0)  # 第一行 值
    print("第一行的值", row_value)
    col_value = table1.col_values(0)  # 列值
    # print(col_value)
    # cell_value = table1.cell_value(0, 5)  # 第一行 第五列
    # print(cell_value)
    alist = []
    # 编列身份证前6位数
    for i in range(1,len(col_value)):
        # print(i,col_value[i])
        alist.append(col_value[i])
        # pass
    # alist.append(i)
    # a=alist.index(2)
    # print(alist)  # 装进集合的所有证件号前六位
    # print(col_value[i])
    print(random.choice(alist))  # 获取到随机的6位数
    return random.choice(alist)


# def test_interface_with_for_loop(path):
#     # 加载 Excel 文件路径
#     wb = load_workbook(path)
#     ws = wb['Sheet1']  # 替换为你的工作表名称
#     # 获取所有测试用例数据
#     max_row = ws.max_row
#     for row in range(2, max_row + 1):  # 跳过标题行（第1行）
#             # 地址
#             ip = ws.cell(row=row, column=1).value,
#             # 接口
#             method = ws.cell(row=row, column=2).value,
#             # 头信息
#             headers = {'Content-Type': 'application/json'},
#             # 请求体
#             data = ws.cell(row=row, column=3).value,
#             # 响应结果
#             # response = ws.cell(row=row, column=4).value
#             print(ip, method, headers, data)

#  获取文件所有的值
def get_data_from_excel(path):
    wb = load_workbook(path)
    ws = wb['Sheet1']  # 替换为你的工作表名称

    data = []
    for row in range(2, ws.max_row + 1):
        test_case = {
            'ip': ws.cell(row=row, column=1).value, # 地址
            'interfase': ws.cell(row=row, column=2).value,  # 接口
            'headers': {'Content-Type': 'application/json'},  # 头
            'data': ws.cell(row=row, column=3).value,   # 请求体
            # 'expected_status_code': int(ws.cell(row=row, column=5).value)  # 期望值
        }
        data.append(test_case)
    # print(data)
    return data


# 遍历每一个值
# def list_read_Excel_value(path):
#     for data in get_data_from_excel(path):
#         ip = data['ip']
#         interfase = data['interfase']
#         headers = data['headers']
#         datas = data['data']
#         # print(data['ip'], data['interfase'], data['headers'], data['data'])
#         print(ip, interfase, headers, datas)


if __name__ =='__main__':
    # a = test_interface_with_for_loop('case.xlsx')
    # print(*a[0], *a[1], *a[2], *a[3])
    print(get_data_from_excel('case.xlsx')[0]['interfase'])
    # print(get_data_from_excel('case.xlsx'))
    # print(list_read_Excel_value('case.xlsx'))

    ...
